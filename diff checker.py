from difflib import SequenceMatcher

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

DEFAULT_FILE1 = r""
DEFAULT_FILE2 = r""
DEFAULT_OUTPUT_FILE = r"C:\Users\suraj\OneDrive\Desktop\comparison_report.xlsx"
DEFAULT_IGNORE_COLUMNS = ["created_timestamp"]
DEFAULT_ROW_MATCH_THRESHOLD = 0.65

# ---------------------------
# Read Files
# ---------------------------

def ask_with_default(prompt, default):
    value = input(f"{prompt} [{default}]: ").strip()
    return value or default

def ask_list_with_default(prompt, default):
    default_text = ", ".join(default)
    value = input(f"{prompt} [{default_text}]: ").strip()

    if not value:
        return default

    return [
        item.strip()
        for item in value.split(",")
        if item.strip()
    ]

def ask_float_with_default(prompt, default):
    while True:
        value = input(f"{prompt} [{default}]: ").strip()

        if not value:
            return default

        try:
            return float(value)
        except ValueError:
            print("Please enter a number, for example 0.65.")

def read_file(file):
    if file.lower().endswith(".csv"):
        return pd.read_csv(file, dtype=str)
    return pd.read_excel(file, dtype=str)

def compare_by_similarity(df1, df2, all_cols, row_match_threshold):
    old_rows = [
        df1.loc[row, all_cols].astype(str).tolist()
        for row in df1.index
    ]
    new_rows = [
        df2.loc[row, all_cols].astype(str).tolist()
        for row in df2.index
    ]
    old_signatures = ["|".join(row) for row in old_rows]
    new_signatures = ["|".join(row) for row in new_rows]

    matcher = SequenceMatcher(
        a=old_signatures,
        b=new_signatures,
        autojunk=False
    )

    output_rows = []
    modified_cells = []
    added_rows = []
    deleted_rows = []

    def row_similarity(old_index, new_index):
        old_text = old_signatures[old_index]
        new_text = new_signatures[new_index]
        return SequenceMatcher(None, old_text, new_text, autojunk=False).ratio()

    def append_modified_or_unchanged_row(old_index, new_index):
        row = len(output_rows)
        output_row = df2.loc[new_index, all_cols].copy()

        for col in all_cols:
            val1 = str(df1.at[old_index, col])
            val2 = str(df2.at[new_index, col])

            if val1 != val2:
                output_row[col] = f"{val1} -> {val2}"
                modified_cells.append((row, col))

        output_rows.append(output_row.to_dict())

    def append_deleted_row(old_index):
        deleted_rows.append(len(output_rows))
        output_rows.append(df1.loc[old_index, all_cols].to_dict())

    def append_added_row(new_index):
        added_rows.append(len(output_rows))
        output_rows.append(df2.loc[new_index, all_cols].to_dict())

    def append_replace_block(old_start, old_end, new_start, new_end):
        old_index = old_start
        new_index = new_start

        while old_index < old_end and new_index < new_end:
            current_similarity = row_similarity(old_index, new_index)

            next_new_similarity = -1
            if new_index + 1 < new_end:
                next_new_similarity = row_similarity(old_index, new_index + 1)

            next_old_similarity = -1
            if old_index + 1 < old_end:
                next_old_similarity = row_similarity(old_index + 1, new_index)

            if current_similarity >= row_match_threshold:
                append_modified_or_unchanged_row(old_index, new_index)
                old_index += 1
                new_index += 1
            elif next_new_similarity >= row_match_threshold and next_new_similarity > next_old_similarity:
                append_added_row(new_index)
                new_index += 1
            elif next_old_similarity >= row_match_threshold:
                append_deleted_row(old_index)
                old_index += 1
            else:
                append_deleted_row(old_index)
                append_added_row(new_index)
                old_index += 1
                new_index += 1

        while old_index < old_end:
            append_deleted_row(old_index)
            old_index += 1

        while new_index < new_end:
            append_added_row(new_index)
            new_index += 1

    for tag, old_start, old_end, new_start, new_end in matcher.get_opcodes():
        if tag == "equal":
            for offset in range(old_end - old_start):
                append_modified_or_unchanged_row(
                    old_start + offset,
                    new_start + offset
                )
        elif tag == "delete":
            for old_index in range(old_start, old_end):
                append_deleted_row(old_index)
        elif tag == "insert":
            for new_index in range(new_start, new_end):
                append_added_row(new_index)
        elif tag == "replace":
            append_replace_block(old_start, old_end, new_start, new_end)

    output_df = pd.DataFrame(output_rows, columns=all_cols)
    return output_df, modified_cells, added_rows, deleted_rows

file1 = ask_with_default("Old/prod file path", DEFAULT_FILE1)
file2 = ask_with_default("New/dev file path", DEFAULT_FILE2)
output_file = ask_with_default("Output Excel file path", DEFAULT_OUTPUT_FILE)
ignore_columns = ask_list_with_default("Columns to ignore, comma-separated", DEFAULT_IGNORE_COLUMNS)
row_match_threshold = ask_float_with_default(
    "Row match threshold from 0 to 1",
    DEFAULT_ROW_MATCH_THRESHOLD
)

df1 = read_file(file1).fillna("")
df2 = read_file(file2).fillna("")

# Clean column names so headers with accidental spaces still match.
df1.columns = df1.columns.str.strip()
df2.columns = df2.columns.str.strip()

# Remove ignored columns
df1 = df1.drop(columns=ignore_columns, errors="ignore")
df2 = df2.drop(columns=ignore_columns, errors="ignore")

# Make columns identical while preserving the report column order.
# Start with df2 because the report is based on df2, then append df1-only columns.
all_cols = list(df2.columns)
all_cols.extend(col for col in df1.columns if col not in all_cols)

df1 = df1.reindex(columns=all_cols, fill_value="")
df2 = df2.reindex(columns=all_cols, fill_value="")

# ---------------------------
# Compare
# ---------------------------

print("Matching rows by sequence and row similarity.")
output_df, modified_cells, added_rows, deleted_rows = compare_by_similarity(
    df1,
    df2,
    all_cols,
    row_match_threshold
)

# ---------------------------
# Save Excel
# ---------------------------

output_df.to_excel(
    output_file,
    sheet_name="Comparison",
    index=False
)

# ---------------------------
# Highlighting
# ---------------------------

wb = load_workbook(output_file)
ws = wb["Comparison"]

yellow = PatternFill(
    fill_type="solid",
    start_color="FFEB9C"
)

green = PatternFill(
    fill_type="solid",
    start_color="C6EFCE"
)

red = PatternFill(
    fill_type="solid",
    start_color="FFC7CE"
)

col_map = {
    col: idx + 1
    for idx, col in enumerate(output_df.columns)
}

# Modified cells
for row, col in modified_cells:
    excel_row = row + 2
    excel_col = col_map[col]

    ws.cell(
        excel_row,
        excel_col
    ).fill = yellow

# Added rows
for row in added_rows:
    excel_row = row + 2

    for col in range(1, ws.max_column + 1):
        ws.cell(excel_row, col).fill = green

# Deleted rows
for row in deleted_rows:
    excel_row = row + 2

    for col in range(1, ws.max_column + 1):
        ws.cell(excel_row, col).fill = red

ws.freeze_panes = "A2"

wb.save(output_file)

print(f"Generated: {output_file}")
