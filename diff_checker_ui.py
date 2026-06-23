"""
UI version of diff checker.py
Keeps the original comparison logic and adds:
  - Drag-and-drop file selection for the two Excel/CSV files
  - Toggle checkboxes to include/exclude columns from comparison
  - Output path picker, threshold input, status feedback, and summary
"""

import os
import shutil
import sys
import tempfile
import threading
import tkinter as tk
from difflib import SequenceMatcher
from tkinter import filedialog, messagebox, ttk

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# Try to use tkinterdnd2 for drag-and-drop. If unavailable, fall back to
# plain tkinter (file dialogs still work, drag-and-drop is disabled).
try:
    from tkinterdnd2 import TkinterDnD, DND_FILES
    HAS_DND = True
except ImportError:  # pragma: no cover
    HAS_DND = False
    TkinterDnD = None  # type: ignore

DEFAULT_OUTPUT_FILE = r"C:\Users\suraj\OneDrive\Desktop\comparison_report.xlsx"
DEFAULT_IGNORE_COLUMNS = ["created_timestamp"]
DEFAULT_ROW_MATCH_THRESHOLD = 0.65

# ---------------------------
# Core comparison logic (preserved from diff checker.py)
# ---------------------------


def read_file(file):
    if file.lower().endswith(".csv"):
        return pd.read_csv(file, dtype=str)
    return pd.read_excel(file, dtype=str)


def compare_by_similarity(df1, df2, all_cols, row_match_threshold):
    old_rows = [df1.loc[row, all_cols].astype(str).tolist() for row in df1.index]
    new_rows = [df2.loc[row, all_cols].astype(str).tolist() for row in df2.index]
    old_signatures = ["|".join(row) for row in old_rows]
    new_signatures = ["|".join(row) for row in new_rows]

    matcher = SequenceMatcher(a=old_signatures, b=new_signatures, autojunk=False)

    output_rows = []
    modified_cells = []
    added_rows = []
    deleted_rows = []

    def row_similarity(old_index, new_index):
        old_text = old_signatures[old_index]
        new_text = new_signatures[new_index]
        return SequenceMatcher(None, old_text, new_text, autojunk=False).ratio()

    def append_modified_or_unchanged_row(old_index, new_index):
        row = len(output_rows)
        output_row = df2.loc[new_index, all_cols].copy()

        for col in all_cols:
            val1 = str(df1.at[old_index, col])
            val2 = str(df2.at[new_index, col])

            if val1 != val2:
                output_row[col] = f"{val1} -> {val2}"
                modified_cells.append((row, col))

        output_rows.append(output_row.to_dict())

    def append_deleted_row(old_index):
        deleted_rows.append(len(output_rows))
        output_rows.append(df1.loc[old_index, all_cols].to_dict())

    def append_added_row(new_index):
        added_rows.append(len(output_rows))
        output_rows.append(df2.loc[new_index, all_cols].to_dict())

    def append_replace_block(old_start, old_end, new_start, new_end):
        old_index = old_start
        new_index = new_start

        while old_index < old_end and new_index < new_end:
            current_similarity = row_similarity(old_index, new_index)

            next_new_similarity = -1
            if new_index + 1 < new_end:
                next_new_similarity = row_similarity(old_index, new_index + 1)

            next_old_similarity = -1
            if old_index + 1 < old_end:
                next_old_similarity = row_similarity(old_index + 1, new_index)

            if current_similarity >= row_match_threshold:
                append_modified_or_unchanged_row(old_index, new_index)
                old_index += 1
                new_index += 1
            elif next_new_similarity >= row_match_threshold and next_new_similarity > next_old_similarity:
                append_added_row(new_index)
                new_index += 1
            elif next_old_similarity >= row_match_threshold:
                append_deleted_row(old_index)
                old_index += 1
            else:
                append_deleted_row(old_index)
                append_added_row(new_index)
                old_index += 1
                new_index += 1

        while old_index < old_end:
            append_deleted_row(old_index)
            old_index += 1

        while new_index < new_end:
            append_added_row(new_index)
            new_index += 1

    for tag, old_start, old_end, new_start, new_end in matcher.get_opcodes():
        if tag == "equal":
            for offset in range(old_end - old_start):
                append_modified_or_unchanged_row(old_start + offset, new_start + offset)
        elif tag == "delete":
            for old_index in range(old_start, old_end):
                append_deleted_row(old_index)
        elif tag == "insert":
            for new_index in range(new_start, new_end):
                append_added_row(new_index)
        elif tag == "replace":
            append_replace_block(old_start, old_end, new_start, new_end)

    output_df = pd.DataFrame(output_rows, columns=all_cols)
    return output_df, modified_cells, added_rows, deleted_rows


# ---------------------------
# UI
# ---------------------------


class DiffCheckerUI:
    def __init__(self, root):
        self.root = root
        self.root.title("Excel Diff Checker")
        self.root.geometry("900x700")
        self.root.minsize(800, 600)

        self.file1_path = ""
        self.file2_path = ""
        self.df1 = None
        self.df2 = None
        self.column_vars = {}

        style = ttk.Style()
        style.configure("TFrame", background="#f5f5f5")
        style.configure("TLabel", background="#f5f5f5", font=("Segoe UI", 10))
        style.configure("TButton", font=("Segoe UI", 10))
        style.configure("Header.TLabel", font=("Segoe UI", 12, "bold"))
        style.configure("Drop.TLabel", background="#ffffff", foreground="#555555",
                        font=("Segoe UI", 10), anchor="center", relief="solid", borderwidth=1)

        main_frame = ttk.Frame(root, padding=15)
        main_frame.pack(fill="both", expand=True)

        # --- File drop zones ---
        file_frame = ttk.LabelFrame(main_frame, text="Files to compare", padding=10)
        file_frame.pack(fill="x", pady=(0, 10))

        self.file1_lbl = ttk.Label(
            file_frame,
            text="Drop OLD / PROD file here\nor click to browse",
            style="Drop.TLabel",
            wraplength=350,
            justify="center"
        )
        self.file1_lbl.grid(row=0, column=0, padx=5, pady=5, sticky="nsew")
        self.file1_lbl.bind("<Button-1>", lambda e: self.browse_file1())

        self.file2_lbl = ttk.Label(
            file_frame,
            text="Drop NEW / DEV file here\nor click to browse",
            style="Drop.TLabel",
            wraplength=350,
            justify="center"
        )
        self.file2_lbl.grid(row=0, column=1, padx=5, pady=5, sticky="nsew")
        self.file2_lbl.bind("<Button-1>", lambda e: self.browse_file2())

        if HAS_DND:
            self.file1_lbl.drop_target_register(DND_FILES)
            self.file1_lbl.dnd_bind("<<Drop>>", self.on_drop_file1)
            self.file2_lbl.drop_target_register(DND_FILES)
            self.file2_lbl.dnd_bind("<<Drop>>", self.on_drop_file2)

        file_frame.columnconfigure(0, weight=1)
        file_frame.columnconfigure(1, weight=1)
        file_frame.rowconfigure(0, minsize=80)

        # --- Options ---
        options_frame = ttk.LabelFrame(main_frame, text="Comparison options", padding=10)
        options_frame.pack(fill="x", pady=(0, 10))

        ttk.Label(options_frame, text="Output file:").grid(row=0, column=0, sticky="w", padx=5, pady=5)
        self.output_entry = ttk.Entry(options_frame)
        self.output_entry.insert(0, DEFAULT_OUTPUT_FILE)
        self.output_entry.grid(row=0, column=1, sticky="ew", padx=5, pady=5)
        ttk.Button(options_frame, text="Browse", command=self.browse_output).grid(row=0, column=2, padx=5, pady=5)

        ttk.Label(options_frame, text="Row match threshold (0-1):").grid(row=1, column=0, sticky="w", padx=5, pady=5)
        self.threshold_entry = ttk.Entry(options_frame, width=10)
        self.threshold_entry.insert(0, str(DEFAULT_ROW_MATCH_THRESHOLD))
        self.threshold_entry.grid(row=1, column=1, sticky="w", padx=5, pady=5)

        options_frame.columnconfigure(1, weight=1)

        # --- Column selection ---
        columns_frame = ttk.LabelFrame(main_frame, text="Columns to include in comparison", padding=10)
        columns_frame.pack(fill="both", expand=True, pady=(0, 10))

        btn_frame = ttk.Frame(columns_frame)
        btn_frame.pack(fill="x", pady=(0, 8))

        ttk.Button(btn_frame, text="Load columns", command=self.load_columns).pack(side="left", padx=(0, 5))
        ttk.Button(btn_frame, text="Select all", command=self.select_all_columns).pack(side="left", padx=5)
        ttk.Button(btn_frame, text="Deselect all", command=self.deselect_all_columns).pack(side="left", padx=5)
        ttk.Button(btn_frame, text="Invert", command=self.invert_columns).pack(side="left", padx=5)

        self.columns_canvas = tk.Canvas(columns_frame, background="#ffffff", highlightthickness=0)
        self.columns_canvas.pack(side="left", fill="both", expand=True)

        scrollbar = ttk.Scrollbar(columns_frame, orient="vertical", command=self.columns_canvas.yview)
        scrollbar.pack(side="right", fill="y")
        self.columns_canvas.configure(yscrollcommand=scrollbar.set)

        self.columns_inner = ttk.Frame(self.columns_canvas)
        self.columns_window = self.columns_canvas.create_window((0, 0), window=self.columns_inner, anchor="nw")

        self.columns_inner.bind("<Configure>", lambda e: self.columns_canvas.configure(scrollregion=self.columns_canvas.bbox("all")))
        self.columns_canvas.bind("<Configure>", lambda e: self.columns_canvas.itemconfig(self.columns_window, width=e.width))

        # --- Compare button ---
        self.compare_btn = ttk.Button(main_frame, text="Compare & Generate Report", command=self.start_comparison)
        self.compare_btn.pack(fill="x", pady=(0, 5))

        # --- Status / summary ---
        self.status_lbl = ttk.Label(main_frame, text="Ready. Drop two Excel/CSV files or click Browse.", foreground="#555555")
        self.status_lbl.pack(fill="x")

        self.summary_lbl = ttk.Label(main_frame, text="", foreground="#333333", justify="left")
        self.summary_lbl.pack(fill="x", pady=(5, 0))

        self.actions_frame = ttk.Frame(main_frame)
        self.actions_frame.pack(fill="x", pady=(5, 0))
        self.open_folder_btn = ttk.Button(
            self.actions_frame,
            text="Open output folder",
            command=self.open_output_folder,
            state="disabled"
        )
        self.open_folder_btn.pack(side="left")
        self._last_output_path = ""

    # ---------------------------
    # Helpers
    # ---------------------------

    def open_output_folder(self):
        if not self._last_output_path or not os.path.exists(self._last_output_path):
            self.set_status("Output file not found.", is_error=True)
            return
        folder = os.path.dirname(os.path.abspath(self._last_output_path))
        try:
            if sys.platform == "win32":
                os.startfile(folder)
            elif sys.platform == "darwin":
                os.system(f'open "{folder}"')
            else:
                os.system(f'xdg-open "{folder}"')
        except Exception as e:
            self.set_status(f"Could not open folder: {e}", is_error=True)

    def set_status(self, text, is_error=False):
        color = "#c00000" if is_error else "#555555"
        self.status_lbl.config(text=text, foreground=color)
        self.root.update_idletasks()

    def update_summary(self, modified=0, added=0, deleted=0, output_path=""):
        lines = [
            f"Modified rows: {modified}",
            f"Added rows:    {added}",
            f"Deleted rows:  {deleted}",
        ]
        if output_path:
            lines.append(f"Output: {output_path}")
        self.summary_lbl.config(text="\n".join(lines))

    def _normalize_dropped_path(self, path):
        # tkinterdnd2 may wrap paths in braces and include multiple files.
        path = path.strip()
        if path.startswith("{") and path.endswith("}"):
            path = path[1:-1]
        return path

    def _is_valid_file(self, path):
        return os.path.isfile(path) and path.lower().endswith((".xlsx", ".xls", ".csv"))

    def _load_dataframe(self, path, label):
        try:
            df = read_file(path).fillna("")
            df.columns = df.columns.str.strip()
            label.config(text=os.path.basename(path), foreground="#0078d7")
            return df
        except Exception as e:
            self.set_status(f"Error reading {os.path.basename(path)}: {e}", is_error=True)
            return None

    # ---------------------------
    # File selection
    # ---------------------------

    def on_drop_file1(self, event):
        path = self._normalize_dropped_path(event.data)
        if self._is_valid_file(path):
            self.file1_path = path
            self.df1 = self._load_dataframe(path, self.file1_lbl)
            self.set_status(f"Loaded old/prod file: {os.path.basename(path)}")
            self.auto_load_columns()
        else:
            self.set_status("Please drop a valid Excel or CSV file.", is_error=True)

    def on_drop_file2(self, event):
        path = self._normalize_dropped_path(event.data)
        if self._is_valid_file(path):
            self.file2_path = path
            self.df2 = self._load_dataframe(path, self.file2_lbl)
            self.set_status(f"Loaded new/dev file: {os.path.basename(path)}")
            self.auto_load_columns()
        else:
            self.set_status("Please drop a valid Excel or CSV file.", is_error=True)

    def browse_file1(self):
        path = filedialog.askopenfilename(filetypes=[("Excel/CSV files", "*.xlsx *.xls *.csv")])
        if path:
            self.file1_path = path
            self.df1 = self._load_dataframe(path, self.file1_lbl)
            self.set_status(f"Loaded old/prod file: {os.path.basename(path)}")
            self.auto_load_columns()

    def browse_file2(self):
        path = filedialog.askopenfilename(filetypes=[("Excel/CSV files", "*.xlsx *.xls *.csv")])
        if path:
            self.file2_path = path
            self.df2 = self._load_dataframe(path, self.file2_lbl)
            self.set_status(f"Loaded new/dev file: {os.path.basename(path)}")
            self.auto_load_columns()

    def browse_output(self):
        path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx")]
        )
        if path:
            self.output_entry.delete(0, "end")
            self.output_entry.insert(0, path)

    # ---------------------------
    # Column toggles
    # ---------------------------

    def auto_load_columns(self):
        if self.df1 is not None and self.df2 is not None:
            self.load_columns()

    def load_columns(self):
        if self.df1 is None or self.df2 is None:
            self.set_status("Load both files first.", is_error=True)
            return

        # Clear existing checkboxes
        for widget in self.columns_inner.winfo_children():
            widget.destroy()
        self.column_vars.clear()

        all_cols = list(self.df2.columns)
        for col in self.df1.columns:
            if col not in all_cols:
                all_cols.append(col)

        if not all_cols:
            ttk.Label(self.columns_inner, text="No columns found.", foreground="#888888").pack(anchor="w", padx=5, pady=2)
            return

        for col in all_cols:
            var = tk.BooleanVar(value=col not in DEFAULT_IGNORE_COLUMNS)
            chk = ttk.Checkbutton(self.columns_inner, text=col, variable=var)
            chk.pack(anchor="w", padx=5, pady=2)
            self.column_vars[col] = var

        self.set_status(f"Loaded {len(all_cols)} columns. Uncheck any columns to ignore.")
        self.columns_canvas.update_idletasks()
        self.columns_canvas.configure(scrollregion=self.columns_canvas.bbox("all"))

    def select_all_columns(self):
        for var in self.column_vars.values():
            var.set(True)

    def deselect_all_columns(self):
        for var in self.column_vars.values():
            var.set(False)

    def invert_columns(self):
        for var in self.column_vars.values():
            var.set(not var.get())

    def get_included_columns(self):
        return [col for col, var in self.column_vars.items() if var.get()]

    # ---------------------------
    # Comparison
    # ---------------------------

    def start_comparison(self):
        if self.df1 is None or self.df2 is None:
            self.set_status("Please load both files first.", is_error=True)
            return

        included = self.get_included_columns()
        if not included:
            self.set_status("Select at least one column to compare.", is_error=True)
            return

        try:
            threshold = float(self.threshold_entry.get())
            if not 0 <= threshold <= 1:
                raise ValueError
        except ValueError:
            self.set_status("Threshold must be a number between 0 and 1.", is_error=True)
            return

        output_path = self.output_entry.get().strip()
        if not output_path:
            self.set_status("Please specify an output file path.", is_error=True)
            return

        self.compare_btn.config(state="disabled")
        self.set_status("Comparing... please wait.")

        thread = threading.Thread(
            target=self._compare_worker,
            args=(self.df1.copy(), self.df2.copy(), included, threshold, output_path),
            daemon=True
        )
        thread.start()

    def _compare_worker(self, df1, df2, included_cols, threshold, output_path):
        try:
            # Make sure the destination folder exists before writing
            output_dir = os.path.dirname(output_path)
            if output_dir and not os.path.exists(output_dir):
                os.makedirs(output_dir)

            # Build the final column list (df2 order first, then df1-only columns)
            all_cols = [col for col in df2.columns if col in included_cols]
            for col in df1.columns:
                if col in included_cols and col not in all_cols:
                    all_cols.append(col)

            # Reindex so both frames have the same columns in the same order
            df1 = df1.reindex(columns=all_cols, fill_value="")
            df2 = df2.reindex(columns=all_cols, fill_value="")

            output_df, modified_cells, added_rows, deleted_rows = compare_by_similarity(
                df1, df2, all_cols, threshold
            )

            # Write to a temporary file first so an open/locked target file
            # cannot leave a partially-written report behind.
            fd, temp_path = tempfile.mkstemp(
                suffix=".xlsx",
                dir=output_dir or ".",
                prefix="comparison_report_"
            )
            os.close(fd)

            try:
                output_df.to_excel(temp_path, sheet_name="Comparison", index=False)

                wb = load_workbook(temp_path)
                ws = wb["Comparison"]

                yellow = PatternFill(fill_type="solid", start_color="FFEB9C")
                green = PatternFill(fill_type="solid", start_color="C6EFCE")
                red = PatternFill(fill_type="solid", start_color="FFC7CE")

                col_map = {col: idx + 1 for idx, col in enumerate(output_df.columns)}

                for row, col in modified_cells:
                    ws.cell(row + 2, col_map[col]).fill = yellow

                for row in added_rows:
                    for col in range(1, ws.max_column + 1):
                        ws.cell(row + 2, col).fill = green

                for row in deleted_rows:
                    for col in range(1, ws.max_column + 1):
                        ws.cell(row + 2, col).fill = red

                ws.freeze_panes = "A2"
                wb.save(temp_path)

                # Move the completed report to the requested path
                shutil.move(temp_path, output_path)
            except Exception:
                # Clean up the temporary file on any error
                if os.path.exists(temp_path):
                    os.remove(temp_path)
                raise

            # Count distinct rows that contain at least one modified cell
            modified_rows = len({row for row, _ in modified_cells})

            self.root.after(0, lambda: self._on_comparison_success(
                output_path, modified_rows, len(added_rows), len(deleted_rows)
            ))
        except Exception as e:
            self.root.after(0, lambda: self._on_comparison_error(str(e)))

    def _on_comparison_success(self, output_path, modified, added, deleted):
        self.compare_btn.config(state="normal")
        self._last_output_path = output_path
        self.open_folder_btn.config(state="normal")
        self.set_status(f"Report generated: {output_path}")
        self.update_summary(modified=modified, added=added, deleted=deleted, output_path=output_path)
        messagebox.showinfo("Comparison complete", f"Report saved to:\n{output_path}")

    def _on_comparison_error(self, error_msg):
        self.compare_btn.config(state="normal")
        self.open_folder_btn.config(state="disabled")
        self._last_output_path = ""
        self.set_status(f"Error: {error_msg}", is_error=True)
        messagebox.showerror("Comparison failed", error_msg)


if __name__ == "__main__":
    if HAS_DND:
        root = TkinterDnD.Tk()
    else:
        root = tk.Tk()
        messagebox.showwarning(
            "Drag-and-drop disabled",
            "tkinterdnd2 is not installed.\n"
            "Install it (e.g. pip install tkinterdnd2) to enable drag-and-drop.\n"
            "You can still use the Browse buttons to select files."
        )

    app = DiffCheckerUI(root)
    root.mainloop()
